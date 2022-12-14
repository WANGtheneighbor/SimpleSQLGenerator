from openpyxl import load_workbook
import os.path
pwd = os.path.dirname(__file__)
book = load_workbook(pwd+'\数据示例.xlsx', data_only=True)

IS_MYSQL = False
res = []
for i, sheet in enumerate(book.sheetnames):
    sql = f'''insert into {sheet}
( '''
    cols = [c for c in next(book[sheet].values)]
    cols = list(filter(None, cols))
    sql += ', '.join(cols)
    sql += ' )\nvalues\n'

    for j, row in enumerate(book[sheet].values):
        if j > 0 and any(row):
            line = '('

            for k, cell in enumerate(row):
                if k >= len(cols):
                    break
                line += '\''+str(cell)+'\'' if cell != None else 'NULL'
                line += ', ' if k < len(cols)-1 else ''
            line += '),\n'
            sql += line
    res.append(sql[:-2]+';\n')

# print(res)
with open(pwd+'\outputdata.sql', 'w+', encoding='utf8') as f:
    f.write('\n'.join(s for s in res))

print("\n\n6\n\n")
